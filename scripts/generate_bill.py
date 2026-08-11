#!/usr/bin/env python
"""
Bill Generator CLI Tool

Generates a styled Excel bill/invoice file from JSON input data.
Uses Decimal for money-safe calculations with ROUND_HALF_UP rounding.
Writes formatted workbook with items table, headers, borders, and totals.

Usage:
    python generate_bill.py --input <path_to_bill.json> --output <output.xlsx>

Example bill.json:
    {
        "bill_number": "B-2026-0001",
        "date": "2026-08-11",
        "customer": "Acme Corp",
        "currency": "$",
        "tax_rate": 8,
        "items": [
            {"name": "Web design", "quantity": 1, "unit_price": 1200.00},
            {"name": "Hosting (monthly)", "quantity": 12, "unit_price": 9.99}
        ]
    }
"""

import argparse
import json
import os
import sys
from datetime import date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font


def quantize(value: Decimal, places: int = 2) -> Decimal:
    """Quantize a Decimal to specified decimal places using ROUND_HALF_UP."""
    quantize_str = '0.' + '0' * places
    return value.quantize(Decimal(quantize_str), rounding=ROUND_HALF_UP)


def compute_totals(items: list, tax_rate_percent: Decimal) -> dict:
    """
    Compute subtotal, tax, and total for a list of bill items.
    
    Args:
        items: List of dicts with 'name', 'quantity', 'unit_price' keys.
        tax_rate_percent: Tax rate as a percentage (e.g., 8 for 8%).
    
    Returns:
        dict with keys 'subtotal', 'tax', 'total' (all Decimal, quantized to 2dp).
    
    Formula:
        subtotal = sum(quantity * unit_price for all items)
        tax = subtotal * (tax_rate_percent / 100)
        total = subtotal + tax
    """
    if not items:
        return {
            "subtotal": quantize(Decimal("0.00")),
            "tax": quantize(Decimal("0.00")),
            "total": quantize(Decimal("0.00")),
        }
    
    # Compute subtotal as sum of (quantity * unit_price)
    subtotal = Decimal("0")
    for item in items:
        quantity = Decimal(str(item["quantity"]))
        unit_price = Decimal(str(item["unit_price"]))
        amount = quantity * unit_price
        subtotal += amount
    
    # Quantize subtotal to avoid float artifacts
    subtotal = quantize(subtotal)
    
    # Compute tax
    tax_rate = Decimal(str(tax_rate_percent)) / Decimal("100")
    tax = subtotal * tax_rate
    tax = quantize(tax)
    
    # Compute total
    total = subtotal + tax
    total = quantize(total)
    
    return {
        "subtotal": subtotal,
        "tax": tax,
        "total": total,
    }


def validate_bill(data: dict) -> None:
    """
    Validate bill data structure and values.
    
    Args:
        data: Bill data dict with keys: 'items' (list), optionally 'bill_number',
              'date', 'customer', 'currency', 'tax_rate'.
    
    Raises:
        ValueError: If validation fails with a clear message indicating the issue.
    """
    # Check 'items' key exists
    if "items" not in data:
        raise ValueError("missing 'items'")
    
    items = data["items"]
    
    # Check items is a list
    if not isinstance(items, list):
        raise ValueError("'items' must be a list")
    
    for idx, item in enumerate(items):
        # Each item must be a dict
        if not isinstance(item, dict):
            raise ValueError(f"Item at index {idx} must be a dict")
        
        # Check required fields exist
        if "name" not in item:
            raise ValueError(f"Item at index {idx} missing 'name'")
        if "quantity" not in item:
            raise ValueError(f"Item at index {idx} missing 'quantity'")
        if "unit_price" not in item:
            raise ValueError(f"Item at index {idx} missing 'unit_price'")
        
        # Check quantity is numeric and non-negative
        try:
            quantity = Decimal(str(item["quantity"]))
        except (TypeError, ValueError):
            raise ValueError(f"Item '{item.get('name', idx)}' has non-numeric quantity")
        
        if quantity < 0:
            raise ValueError(f"Item '{item['name']}' has negative quantity: {quantity}")
        
        # Check unit_price is numeric and non-negative
        try:
            unit_price = Decimal(str(item["unit_price"]))
        except (TypeError, ValueError, InvalidOperation):
            raise ValueError(f"Item '{item.get('name', idx)}' has non-numeric unit_price")
        
        if unit_price < 0:
            raise ValueError(f"Item '{item['name']}' has negative unit_price: {unit_price}")


def generate_bill_workbook(data: dict) -> Workbook:
    """
    Generate a styled Excel workbook for a bill.
    
    Args:
        data: Bill data dict with keys: 'bill_number', 'date', 'customer',
              'currency', 'tax_rate', 'items'.
    
    Returns:
        An openpyxl Workbook object with a 'Bill' sheet containing the formatted data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill"
    
    # Extract data with defaults
    bill_number = data.get("bill_number", "N/A")
    bill_date = data.get("date")
    if not bill_date:
        bill_date = date.today().isoformat()
    customer = data.get("customer", "N/A")
    currency = data.get("currency", "$")
    tax_rate = Decimal(str(data.get("tax_rate", 0)))
    items = data.get("items", [])
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    border_thin_no_left = Border(
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    total_font = Font(bold=True)
    
    # Helper to apply borders
    def apply_borders(cell):
        if cell.value is not None:
            cell.border = border_thin
        return cell
    
    # Helper to format amount cells (currency)
    def format_amount(cell, amount):
        cell.value = amount
        cell.number_format = '#,##0.00'
        cell.border = border_thin
        return cell
    
    # Helper to create total row with currency label
    def create_total_row(row_num, label, amount):
        cell_label = ws.cell(row=row_num, column=1)
        cell_label.value = f"{label} ({currency})"
        cell_label.font = total_font
        
        cell_amount = ws.cell(row=row_num, column=4)
        cell_amount.value = amount
        cell_amount.number_format = '#,##0.00'
        cell_amount.font = total_font
        cell_amount.border = border_thin
        
        return row_num
    
    # Write top block info
    ws.cell(row=1, column=1, value="Bill Number:")
    ws.cell(row=1, column=2, value=bill_number)
    ws.cell(row=2, column=1, value="Date:")
    ws.cell(row=2, column=2, value=bill_date)
    ws.cell(row=3, column=1, value="Customer:")
    ws.cell(row=3, column=2, value=customer)
    
    # Write header row for items table
    headers = ["Name", "Quantity", "Unit Price", "Amount"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        # Set column widths using column_dimensions
        if col_idx == 1:
            ws.column_dimensions["A"].width = 25  # Name
        elif col_idx == 2:
            ws.column_dimensions["B"].width = 12  # Quantity
        elif col_idx == 3:
            ws.column_dimensions["C"].width = 14  # Unit Price
        else:
            ws.column_dimensions["D"].width = 14  # Amount
    
    # Write item rows
    for idx, item in enumerate(items, start=5):
        name = item.get("name", "")
        quantity = item.get("quantity", 0)
        unit_price = item.get("unit_price", Decimal("0"))
        amount = Decimal(str(quantity)) * Decimal(str(unit_price))
        amount = quantize(amount)
        
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=quantity)
        ws.cell(row=idx, column=3, value=unit_price)
        format_amount(ws.cell(row=idx, column=4), amount)
    
    # Write totals section
    # Determine how many items we had
    num_items = len(items)
    
    # Compute the actual totals
    tax_rate = Decimal(str(data.get("tax_rate", 0)))
    totals = compute_totals(items, tax_rate)
    subtotal = totals["subtotal"]
    tax = totals["tax"]
    total = totals["total"]
    
    # Subtotal row
    subtotal_row = create_total_row(4 + num_items + 1, "Subtotal", subtotal)
    
    # Tax row
    tax_row = create_total_row(4 + num_items + 2, "Tax", tax)
    # Tax label
    ws.cell(row=tax_row, column=2, value=f"({tax_rate}%):")
    
    # Total row
    total_row = create_total_row(4 + num_items + 3, "Total", total)
    
    # Adjust column widths for totals section
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    
    return wb


def main():
    """
    Main CLI entry point.
    
    Returns:
        Exit code: 0 on success, non-zero on error.
    """
    parser = argparse.ArgumentParser(
        description="Generate a bill/invoice Excel file from JSON data."
    )
    parser.add_argument(
        "--input",
        default=os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "scripts", "examples", "bill.json"
        ),
        help="Path to input JSON file containing bill data.",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "bill.xlsx"),
        help="Path to output Excel file.",
    )
    
    args = parser.parse_args()
    
    # Read and parse JSON
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        return 1
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in input file: {e}", file=sys.stderr)
        return 1
    
    # Validate bill data
    try:
        validate_bill(data)
    except ValueError as e:
        print(f"Error: Validation failed: {e}", file=sys.stderr)
        return 1
    
    # Compute totals
    tax_rate = Decimal(str(data.get("tax_rate", 0)))
    totals = compute_totals(data.get("items", []), tax_rate)
    
    # Generate workbook
    try:
        wb = generate_bill_workbook(data)
    except Exception as e:
        print(f"Error: Failed to generate workbook: {e}", file=sys.stderr)
        return 1
    
    # Write to file
    try:
        wb.save(args.output)
        print(f"Bill generated successfully: {args.output}")
        print(f"  Subtotal: {totals['subtotal']}")
        print(f"  Tax ({tax_rate}%): {totals['tax']}")
        print(f"  Total: {totals['total']}")
    except Exception as e:
        print(f"Error: Failed to save workbook: {e}", file=sys.stderr)
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
